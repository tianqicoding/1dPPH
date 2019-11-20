import numpy as np
import torch as th
import csv
import pandas as pd 
import collections
from numpy import matrix, rank
from numpy.linalg import matrix_rank
import networkx as nx

#import matplotlib.pyplot as plt
import sys
import heapq
#from GF import GF
import sympy 
from scipy.linalg import lu
from minBasis import minBasis
from PPH import persHomo
import timeit
from openpyxl import load_workbook
# g=[]
# for i in range(1997, 2016):
# 	with open('data/useconomic/1997.txt', 'r') as f:
# 	    l = [[float(num) for num in line.split(' ')[:-1]] for line in f]
# 	mat=np.matrix(l)
# 	mat.resize((15, 15))
# 	#print(mat.shape)
# 	G=nx.from_numpy_matrix(mat, create_using=nx.DiGraph)
# 	g.append(persHomo(G))
# start = timeit.default_timer()
# for gi in g:
# 	gi.perHom(2)
# stop = timeit.default_timer()
# print('Time: ', stop - start)
# print(len(g.TreeEdge), len(g.curG.edges), len(g.Bnd))
Country=['AF', 'AM', 'AZ', 'BA', 'BD', 'BT', 'BN', 'KH', 'CN', 'HK', 'MO', 'CY', 'KP', 'GE', 'IN', 'ID', 'IR', 'IQ', 'IL', 'JP', 'JO', 'KZ', 'KW', 'KG', 'LA', 'LB', 'MY', 'MV', 'MN', 'MM', 'NP', 'OM', 'PK', 'PH', 'QA', 'KR', 'SA', 'SG', 'LK', 'PS', 'SY', 'TJ', 'TH', 'TI', 'TR', 'TM', 'AE', 'UZ', 'VN', 'YE']
word='migration'
wb = load_workbook(word+'.xlsx', read_only=True)
#print(wb.sheetnames)
ws = wb['Sheet1']
#cols=0
x2 = np.array([[r[cols].value for r in ws.iter_rows()]for cols in range(50)]).transpose()
#print(x2)

print(x2.shape)
for i in range(x2.shape[0]):
	for j in range(x2.shape[1]):
		if x2[i, j]<0:
			x2[j,i]=-x2[i, j]
			x2[i, j]=0
mx=x2.max()
for i in range(x2.shape[0]):
	for j in range(x2.shape[1]):
		if x2[i, j]!=0:
			x2[i,j]=mx+1-x2[i, j]
			
np.savetxt(word+'.csv', x2, delimiter=',')
G=nx.from_numpy_matrix(x2, create_using=nx.DiGraph)

print(len(G.nodes), len(G.edges))
g=persHomo(G)
start = timeit.default_timer()
g.perHom(3487400)
stop = timeit.default_timer()
g.findCycle(word+'_Generators.txt')
file=open(word+'_Generators.txt',"w")
for cycle in g.generatorCycles:
	for e in cycle:
		file.write('(')
		file.write(Country[e[0]])
		file.write('\t')
		file.write(Country[e[1]])
		file.write(')')
		file.write('\t')
	file.write('\n')
file.close()

print(g.generatorCycles)
g.PrintPair()
print('Time: ', stop - start)
# # minB=minBasis(g)
# # minB.ComputeAnnotation()
# # minB.ComputeMinimalBasis('migration')
